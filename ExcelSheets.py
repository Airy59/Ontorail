"""
Excel-to-RDF transformation classes.
"""

import datetime

import openpyxl
from rdflib import RDF, Literal, URIRef

from References import nsRoo, RooGraph
from Utils import to_title, to_name_en, to_name_zh, prepare_for_SMW_import


class DataReqFile:
	"""
	Ancestor Class for IFC Data Requirements files description and processing
	"""

	def __init__(self, path, tab_obj, tab_prop, version):
		self.DictName = 'IFC_2019'
		self.NameSpace = 'IFC:'
		self.Path = path
		self.ImportDateTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
		self.TabObj = tab_obj
		self.TabProp = tab_prop
		self.Book = openpyxl.load_workbook(self.Path, read_only=True)
		self.SheetObj = self.Book[self.TabObj]
		self.SheetProp = self.Book[self.TabProp]
		self.Version = version
		self.Rdf = None
		self.RdfFormat = None
		self.DomainPackages = {'Sig': 'Signalling Package', 'Tra': 'Track Package', 'Tel': 'Telecom Package',
		                       'Ene': 'Energy Package'}
		# Triples
		self.Graph = RooGraph(identifier='IFC_Data_Req_')
		for dp in self.DomainPackages.values():
			self.Graph.add((to_title(dp), RDF.type, Literal("Domain_Package")))
			self.Graph.add((to_title(dp), nsRoo.BelongsToDictionary, Literal(self.DictName)))
			self.Graph.add((to_title(dp), nsRoo.HasVersion, Literal(version)))
			self.Graph.add((to_title(dp), nsRoo.HasNameEn, Literal(dp)))
			self.Graph.add((to_title(dp), nsRoo.Wikitext, Literal(
					R"Functional categories in this Domain Package: {{{{#ask: [[Category:Functional Category]] [[BelongsToDictionary::{}]] [[InDomainPackage::{}]] }}}}".format(
							Literal(self.DictName),dp))))

	def get_objects(self):
		pass

	def cast_to_rdf(self, path, fmt='turtle'):
		self.RdfFormat = fmt
		self.Rdf = open(path, 'w+t', newline=None)
		s = str(self.Graph.serialize(format=fmt), 'utf-8')
		s = prepare_for_SMW_import(s)
		outfile = open(path, 'w+t', encoding='utf-8', newline=None)
		outfile.write(s)
		outfile.close()

	def find_multiple_definitions(self):
		"""Find objects or properties with multiple definitions (in one language)"""
		pass


class SIG(DataReqFile):
	"""
	SIG data requirements
	"""

	def __init__(self, path, tab_obj, tab_prop_spec, tab_prop_shared, version):
		super().__init__(path, tab_obj, tab_prop_spec, version)
		self.TabPropShared = tab_prop_shared
		self.SheetPropShared = self.Book[tab_prop_shared]
		self.Functional_Categories_Columns = {7: 'CBI', 8: 'Block system', 9: 'Train control system',
		                                      10: 'Traffic dispatching system'}
		# note that curly braces must be doubled to be escaped, herebelow:
		self.Functional_Categories_FreeText = R"Objects belonging to this category: {{{{#ask: [[Category:Object]] [[BelongsToDictionary::{}]] [[InFunctionalCategory::{}]] | ?HasNameZh }}}}"

	def set_functional_categories(self):
		"""
		In SIG, objects may belong to several functional categories!
		The categories themselves are not read from the sheets;
		however, the assignment of objects to categories is documented in the sheets and imported here
		"""
		for v in self.Functional_Categories_Columns.values():
			this_title = to_title(v)
			self.Graph.add((this_title, nsRoo.BelongsToDictionary, Literal('IFC 2019')))
			self.Graph.add((this_title, RDF.type, Literal('Functional Category')))
			self.Graph.add((this_title, nsRoo.HasNameEn, Literal(to_name_en(v))))
			self.Graph.add((this_title, nsRoo.InDomainPackage, Literal('Signalling Package')))
			self.Graph.add((this_title, nsRoo.Wikitext, Literal(self.Functional_Categories_FreeText.format(self.DictName, v))))

	def get_objects(self, first_row, last_row, suffix=''):
		"""
		Scrutinize row after row in selected range.
		:param first_row: first object row
		:param last_row: last object row
		:param suffix: allows to add suffix to object names, for collision avoidance
		:return:
		"""

		cols = {'obj_id': 1, 'obj_name': 2}  # columns A = 1, etc.

		object_count = 0
		# One object at a time...
		for row in self.SheetObj.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=10):
			this_title = URIRef(to_title(row[1].value) + '_--_' + suffix)
			self.Graph.add((this_title, RDF.type, Literal("Object")))
			self.Graph.add((this_title, nsRoo.BelongsToDictionary, Literal('IFC_2019')))
			self.Graph.add((this_title, nsRoo.HasId, Literal(row[cols['obj_id']-1].value)))
			self.Graph.add((this_title, nsRoo.HasVersion, Literal(self.Version)))
			self.Graph.add((this_title, nsRoo.HasNameEn, Literal(to_name_en(row[cols['obj_name']-1].value))))
			self.Graph.add((this_title, nsRoo.HasNameZh, Literal(to_name_zh(row[cols['obj_name']-1].value))))
			for col, fc in self.Functional_Categories_Columns.items():
				# caution: tuple is 0-based, while columns are 1-based...
				if 'x' in str(row[col - 1].value).lower():
					self.Graph.add((this_title, nsRoo.InFunctionalCategory, Literal(fc)))

			object_count += 1

		return (object_count, len(self.Graph))

	def get_properties(self, first_row, last_row, suffix=''):
		"""Get specific properties"""
		cols = {'prop_id': 1, 'object_id': 2, 'prop_group': 3, 'name_en': 4, 'description_en': 5, 'name_zh': 39, 'description_zh': 40}
		property_count = 0
		for row in self.SheetProp.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=42):
			object_name = row[cols['object_id'] - 1].value
			if object_name not in ('', None):
				this_object = list(self.Graph.subjects(predicate=nsRoo.HasNameEn, object=Literal(object_name)))
				if this_object != []:
					this_object = this_object[0]
					this_property = row[cols['name_en'] - 1].value
					if this_property not in ('', None):
						this_title = to_title(this_property)
						property_count += 1
						self.Graph.add((this_title, RDF.type, Literal("Property")))
						self.Graph.add((this_title, nsRoo.HasId, Literal(row[cols['prop_id']-1].value)))
						self.Graph.add((this_title, nsRoo.BelongsToDictionary, Literal('IFC_2019')))
						self.Graph.add((this_title, nsRoo.CharacterizesObject, this_object))  # w/o "Literal", otherwise excess column
						self.Graph.add((this_title, nsRoo.BelongsToGroup, Literal(row[cols['prop_group']-1].value)))
						self.Graph.add((this_title, nsRoo.HasNameEn, Literal(row[cols['name_en'] - 1].value)))
						self.Graph.add((this_title, nsRoo.HasVersion, Literal(self.Version)))
						self.Graph.add(
								(this_title, nsRoo.HasDefinitionEn, Literal(row[cols['description_en'] - 1].value)))
						self.Graph.add((this_title, nsRoo.HasNameZh, Literal(row[cols['name_zh'] - 1].value)))
						self.Graph.add(
								(this_title, nsRoo.HasDefinitionZh, Literal(row[cols['description_zh'] - 1].value)))
		print('Specific property count (incl. duplicates) : ', property_count)

	def get_shared_properties(self, first_row, last_row):
		"""Get shared properties"""
		cols = {'object_id': 1, 'name_en': 4, 'description_en': 5, 'name_zh': 41, 'description_zh': 42}
		property_count = 0
		for row in self.SheetProp.iter_rows(min_row=first_row, max_row=last_row, min_col=1, max_col=42):
			pass




sig = SIG(R'C:\Users\amagn\Desktop\SIG Data\Copy of 20190606-IFC-SD-005-DataRequirement.xlsx',
          '1-Object_Description ', '2.2-Property_Requirements_Spec', '2.1-Property_Requirement_Shared', "0.1")
result = sig.set_functional_categories()
result = sig.get_objects(6, 46, suffix='sig')
# print('\nTotal number of objects in {}: {}'.format(sig.Graph.n3(), result[0]))
# print('Total number of triples: {}\n'.format(result[1]))
sig.get_properties(12, 1000, suffix='sig')
sig.cast_to_rdf(R'C:\Users\amagn\OneDrive\Dev\DataRequirementsToRDF\SIG.ttl')
